import openpyxl
import os

os.chdir('./document proceed')
wb = openpyxl.load_workbook('./produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

#更改部分产品价格
price_update={
    'Garlic':3.07,
    'Celery':1.19,
    'Lemon':1.27
}

for i in range(2,sheet.max_row+1):
    productName=sheet.cell(row=i,column=1).value
    if productName in price_update:
        sheet.cell(row=i,column=2).value=price_update[productName]
wb.save('updateSale.xlsx')
