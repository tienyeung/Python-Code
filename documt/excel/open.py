import webbrowser
import openpyxl
import urllib

wb = openpyxl.load_workbook('./esi_script.xlsx')
sheet = wb.get_sheet_by_name('Sheet0')   
baseUrl = 'https://cn.bing.com/search?'
for i in range(200,202):
    if sheet.cell(row=i,column=3).value != None:
        continue
    else:
        EsiName=sheet.cell(row=i,column=2).value
        word = EsiName.encode(encoding='utf-8', errors='strict')
        data = {'q':word}
        data = urllib.parse.urlencode(data)
        url = baseUrl+data
        webbrowser.open(url)