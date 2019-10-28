import openpyxl
import os
import requests
import urllib
from bs4 import BeautifulSoup
import re


headers = {"User-Agent":"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.146 Safari/537.36"
 }

wb = openpyxl.load_workbook('./esi_script.xlsx')
sheet = wb.get_sheet_by_name('Sheet0')   

baseUrl = 'https://mijisou.com/?'
for i in range(16,518):
    EsiName=sheet.cell(row=i,column=2).value
    word = EsiName.encode(encoding='utf-8', errors='strict')
    data = {'q':word}
    data = urllib.parse.urlencode(data)
    url = baseUrl+data
    print(url)
    res=requests.get(url=url,headers=headers)
    soup = BeautifulSoup(res.content, 'html.parser')
    item_a=soup.select('h4>a')
    for href in item_a:
        url=href.get('href')
        break
    print(url)
    pat=r"(\w+)(-\w+)+"
    
    result = re.search(pat,url)
    if result:
        result=result.group().replace("-"," ")
    else:
        result=" "
    print(result)
    sheet.cell(row=i,column=3).value=result
wb.save('./UpdateESI.xlsx')











